"""사군자 회비 대시보드  |  python dashboard.py  →  http://localhost:8888"""
import json, os
from http.server import ThreadingHTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', '사군자.xlsx')

def eval_formula(expr):
    if not isinstance(expr, str): return expr
    if expr.startswith('='): expr = expr[1:]
    try: return float(eval(expr.replace('INT(','int('), {'int':int,'abs':abs,'__builtins__':{}}))
    except: return None

def ym_sort_key(ym):
    try:
        p = ym.replace('년 ','-').replace('월','').strip().split('-')
        return (int(p[0]), int(p[1]))
    except: return (99,99)

def load_member_status():
    """입금현황 2행에서 유지/탈퇴 읽기. 미표기는 유지로 간주. 탈퇴만 제외."""
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['입금현황']
    rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    headers, statuses = rows[0], rows[1]
    active = []
    for name, status in zip(headers, statuses):
        if name in (None, 'No', '탈퇴멤버', '계'): continue
        if status == '탈퇴': continue   # 탈퇴만 제외, 미표기는 유지로 간주
        active.append(name)
    return active

def load_income_by_month():
    # 회비 입금은 탈퇴멤버 포함 전체 합산 (실제 계좌 입금액 기준)
    df = pd.read_excel(EXCEL_PATH, sheet_name='입금현황')
    members = [c for c in df.columns if c not in ['No','계'] and not str(c).startswith('Unnamed')]
    data = df[df['No'].apply(lambda x: isinstance(x,str) and '년' in str(x))].copy()
    result = {}
    for _, row in data.iterrows():
        total = sum(pd.to_numeric(row[m],errors='coerce') for m in members if pd.notna(row[m]))
        result[row['No']] = int(total)
    return result

def load_all_payment_status():
    """모든 월의 납부 현황 반환 (유지 멤버만)"""
    active = load_member_status()
    df = pd.read_excel(EXCEL_PATH, sheet_name='입금현황')
    members = [c for c in df.columns if c in active]
    data = df[df['No'].apply(lambda x: isinstance(x,str) and '년' in str(x))]
    real = data[data[members].apply(lambda r: r.notna().any(), axis=1)]
    all_status = {}
    for ym in real['No'].tolist():
        row = df[df['No']==ym].iloc[0]
        all_status[ym] = {m: (pd.notna(row[m]) and float(row[m])>0) for m in members}
    months = list(all_status.keys())
    latest = months[-1] if months else ''
    return all_status, latest, members

def load_wonjang_rows():
    wb = load_workbook(EXCEL_PATH)
    ws = wb['원장']
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=False):
        일자_v=row[2].value; 분류1=row[3].value; 분류2=row[4].value
        차변_r=row[5].value; 대변_r=row[6].value; 비고=row[8].value
        if 일자_v is None or 일자_v=='입력': continue
        try: 일자=pd.to_datetime(일자_v)
        except: continue
        연월=f"{str(일자.year)[2:]}년 {일자.month}월"
        차변=float(eval_formula(차변_r)) if 차변_r is not None and eval_formula(차변_r) is not None else 0.0
        대변=float(eval_formula(대변_r)) if 대변_r is not None and eval_formula(대변_r) is not None else 0.0
        rows.append({'일자':일자,'연월':연월,'분류1':분류1,'분류2':분류2,'차변':차변,'대변':대변,'순액':대변-차변,'비고':str(비고) if 비고 else ''})
    return rows

def load_member_totals():
    # 탈퇴 포함 전체 (잔액 계산용)
    df = pd.read_excel(EXCEL_PATH, sheet_name='입금현황')
    members = [c for c in df.columns if c not in ['No','탈퇴멤버','계'] and not str(c).startswith('Unnamed')]
    data = df[df['No'].apply(lambda x: isinstance(x,str) and '년' in str(x))].copy()
    return [{'name':m,'paid':int(pd.to_numeric(data[m],errors='coerce').fillna(0).sum())} for m in members]

def build_data():
    rows = load_wonjang_rows()
    income_by_month = load_income_by_month()
    all_pay_status, latest_ym, all_members = load_all_payment_status()
    extra_income={}
    for r in rows:
        if r['분류1']=='입금' and r['분류2'] in ('지각비','이자'):
            extra_income[r['연월']]=extra_income.get(r['연월'],0)+r['대변']
    expense_by_cat={}; expense_by_month={}; expense_by_month_meeting={}
    for r in rows:
        if r['분류1']=='출금' and r['차변']>0:
            net = r['차변'] - (r['대변'] if r['대변'] else 0)  # 대변(환불/환급)이 있으면 차감, 음수도 반영
            expense_by_cat[r['분류2']]=expense_by_cat.get(r['분류2'],0)+net
            expense_by_month[r['연월']]=expense_by_month.get(r['연월'],0)+net
            if r['분류2']!='경조사비':  # 월평균 카드용: 경조사비(비경상) 제외한 모임비 기준
                expense_by_month_meeting[r['연월']]=expense_by_month_meeting.get(r['연월'],0)+net
    total_in=sum(income_by_month.values())+sum(extra_income.values())
    total_out=sum(expense_by_cat.values())
    balance=total_in-total_out
    all_months=sorted(set(list(income_by_month)+list(expense_by_month)),key=ym_sort_key)
    monthly=[]; cumulative=0
    for ym in all_months:
        inc=income_by_month.get(ym,0)+extra_income.get(ym,0)
        exp=expense_by_month.get(ym,0)
        cumulative+=inc-exp
        monthly.append({'month':ym,'income':int(inc),'expense':int(exp),'cumulative':int(cumulative)})
    mwi=[m for m in monthly if m['income']>0]
    mwo=[m for m in monthly if m['expense']>0]
    meeting_monthly=[{'month':ym,'expense':int(expense_by_month_meeting.get(ym,0))} for ym in all_months]
    mwo_meeting=[m for m in meeting_monthly if m['expense']>0]
    def stats_n(vals_list, n, key):
        sub=vals_list[-n:] if len(vals_list)>=n else vals_list
        if not sub: return {'mean':0,'median':0}
        vals=sorted(m[key] for m in sub)
        mean=int(sum(vals)/len(vals))
        ln=len(vals)
        median=vals[ln//2] if ln%2==1 else (vals[ln//2-1]+vals[ln//2])/2
        return {'mean':mean,'median':int(round(median))}
    income_stats={str(n):stats_n(mwi,n,'income') for n in (3,6,12)}
    expense_stats={str(n):stats_n(mwo,n,'expense') for n in (3,6,12)}
    meeting_expense_stats={str(n):stats_n(mwo_meeting,n,'expense') for n in (3,6,12)}
    cat_data=[{'name':k,'value':int(v)} for k,v in sorted(expense_by_cat.items(),key=lambda x:-x[1]) if v>0]
    member_totals=load_member_totals()
    valid=[r for r in rows if r['차변']>0 or r['대변']>0]
    for r in valid:
        if r['분류2']=='회비' and r['분류1']=='입금' and r['대변']==0:
            r['대변']=income_by_month.get(r['연월'],0); r['순액']=r['대변']
    recent=sorted(valid,key=lambda r:r['일자'],reverse=True)[:10]
    recent_list=[{'date':r['일자'].strftime('%y년 %m월 %d일'),'type':r['분류2'],'amount':int(r['순액']),'memo':r['비고'],'is_income':r['분류1']=='입금'} for r in recent]
    member_names=[mt['name'] for mt in member_totals]
    payment_by_month={
        ym: [{'name':m,'paid':st.get(m,False)} for m in all_members if m in member_names]
        for ym,st in all_pay_status.items()
    }
    pay_months=list(payment_by_month.keys())
    return {
        'summary':{'total_in':total_in,'total_out':total_out,'balance':balance,'income_stats':income_stats,'expense_stats':expense_stats,'meeting_expense_stats':meeting_expense_stats,'updated':datetime.now().strftime('%Y-%m-%d %H:%M'),'latest_ym':latest_ym},
        'monthly':monthly,'categories':cat_data,'members':member_totals,
        'payment_by_month':payment_by_month,'pay_months':pay_months,'recent':recent_list,
    }

HTML = '''<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>사군자 회비</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=DM+Sans:wght@300;400;500&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
:root{
  --bg:#0E0F12;
  --s1:#16181D;
  --s2:#1B1E24;
  --s3:#20232A;
  --line:rgba(232,232,234,0.07);
  --line2:rgba(232,232,234,0.13);
  --p:#00D4A0;
  --p-dim:rgba(0,212,160,0.10);
  --p-dim2:rgba(0,212,160,0.18);
  --w:#E8E8EA;
  --w2:rgba(232,232,234,0.68);
  --w3:rgba(232,232,234,0.38);
  --w4:rgba(232,232,234,0.16);
  --neg:#ff5e5e;
  --neg-dim:rgba(255,94,94,0.10);
  --fc-pos:#4da3ff;
  --fc-pos-dim:rgba(77,163,255,0.10);
  --fc-pos-dim2:rgba(77,163,255,0.18);
  --fc-neg:#ffb020;
  --fc-neg-dim:rgba(255,176,32,0.10);
  --fc-neg-dim2:rgba(255,176,32,0.18);
}
*{box-sizing:border-box;margin:0;padding:0;}
html{background:var(--bg);}
body{background:var(--bg);color:var(--w);font-family:'DM Sans',sans-serif;font-weight:400;min-height:100vh;-webkit-font-smoothing:antialiased;}

/* NAV */
nav{
  position:sticky;top:0;z-index:50;
  background:rgba(14,15,18,0.82);
  backdrop-filter:blur(24px);
  border-bottom:1px solid var(--line);
  padding:0 28px;
  height:52px;
  display:flex;align-items:center;justify-content:space-between;
}
.nav-logo{font-family:'DM Serif Display',serif;font-size:16px;color:var(--w2);letter-spacing:0.3px;}
.nav-logo em{color:var(--p);font-style:normal;}
.nav-right{display:flex;align-items:center;gap:14px;}
.nav-time{font-size:11px;color:var(--w4);letter-spacing:0.5px;font-variant-numeric:tabular-nums;}
.nav-btn{background:transparent;border:1px solid var(--line2);color:var(--w3);padding:5px 14px;border-radius:6px;cursor:pointer;font-family:inherit;font-size:11px;letter-spacing:0.8px;transition:border-color .2s,color .2s;}
.nav-btn:hover{border-color:var(--p);color:var(--p);}

.page{max-width:720px;margin:0 auto;padding:40px 20px 80px;}

/* TIER 1 — HERO (압도적 주인공) */
.t-hero{padding:8px 0 4px;}
.t-hero-eyebrow{font-size:10px;letter-spacing:3px;color:var(--w4);text-transform:uppercase;margin-bottom:14px;}
.t-hero-num{font-family:'DM Serif Display',serif;font-size:80px;line-height:1;letter-spacing:-3px;color:var(--w);font-variant-numeric:tabular-nums;}
.t-hero-num sup{font-size:28px;letter-spacing:0;vertical-align:super;color:var(--w3);margin-right:3px;}
.t-hero-meta{margin-top:16px;font-size:13px;color:var(--w4);display:flex;align-items:center;gap:9px;}
.t-hero-dot{color:var(--w4);}
#h-net{color:var(--p);font-weight:500;}
#h-net.neg{color:var(--neg);}

/* TIER 2 — 이번달 흐름 (중간) */
.t-month{padding:28px 0 26px;border-top:1px solid var(--line);margin-top:24px;}
.t-month-head{display:flex;align-items:baseline;justify-content:space-between;margin-bottom:18px;}
.t-month-title{font-size:11px;letter-spacing:1.5px;color:var(--w2);text-transform:uppercase;font-weight:500;}
.t-month-period{font-size:11px;color:var(--w4);}
.t-month-row{display:grid;grid-template-columns:repeat(3,1fr);gap:16px;}
.t-month-lbl{font-size:10px;color:var(--w4);letter-spacing:0.6px;text-transform:uppercase;margin-bottom:7px;}
.t-month-val{font-size:28px;font-weight:300;color:var(--w2);letter-spacing:-0.8px;font-variant-numeric:tabular-nums;}
#m-in{color:var(--p);}
.t-month-val.pos{color:var(--p);}
.t-month-val.neg{color:var(--neg);}

/* 핵심 운영 지표 — 월평균 모임비 (두 번째 강조점, 유일한 강조 서페이스) [확정: B안 뉴트럴/화이트] */
.t-focus{
  margin-top:22px;
  padding:24px 26px;
  border-radius:16px;
  background:linear-gradient(180deg, rgba(232,232,234,0.06), rgba(232,232,234,0.015));
  border:1px solid rgba(232,232,234,0.14);
  border-top:2px solid rgba(232,232,234,0.55);
}
.t-focus-head{display:flex;align-items:center;justify-content:space-between;margin-bottom:16px;gap:10px;flex-wrap:wrap;}
.t-focus-titlewrap{display:flex;align-items:baseline;gap:8px;}
.t-focus-title{font-size:12px;letter-spacing:1.3px;color:var(--w);text-transform:uppercase;font-weight:500;}
.t-focus-note{font-size:9px;color:var(--w4);letter-spacing:0.2px;}
.t-focus-val{font-size:42px;font-weight:500;color:var(--w);letter-spacing:-1.3px;font-variant-numeric:tabular-nums;line-height:1;}
.t-focus-sub{font-size:11px;color:var(--w3);margin-top:8px;}
.t-focus-secondary{display:flex;align-items:baseline;gap:9px;padding-top:16px;margin-top:16px;border-top:1px solid rgba(232,232,234,0.08);font-size:11px;color:var(--w4);flex-wrap:wrap;}
.t-focus-sec-lbl{letter-spacing:0.3px;}
.t-focus-sec-val{color:var(--w3);font-weight:500;font-variant-numeric:tabular-nums;}
.t-focus-sec-sub{color:var(--w4);}

/* TIER 3 — 상세 (작게 · 옅게, 박스 대신 헤어라인+여백) */
.t-detail{padding-top:4px;}
.d-block{padding:26px 0;border-top:1px solid var(--line);}
.d-head{display:flex;align-items:center;justify-content:space-between;margin-bottom:18px;gap:10px;flex-wrap:wrap;}
.d-title{font-size:10px;letter-spacing:2px;color:var(--w4);text-transform:uppercase;}
.d-row{display:grid;grid-template-columns:repeat(2,1fr);gap:24px;}
.d-stat{display:flex;flex-direction:column;gap:4px;}
.d-stat-lbl{font-size:10px;letter-spacing:1px;color:var(--w4);text-transform:uppercase;}
.d-stat-val{font-size:19px;font-weight:300;color:var(--w2);letter-spacing:-0.3px;font-variant-numeric:tabular-nums;}
.d-stat-sub{font-size:10px;color:var(--w4);}

/* TABS / SELECT */
.tabs{display:flex;gap:2px;background:var(--s2);padding:3px;border-radius:8px;}
.tab{font-size:11px;padding:4px 12px;border-radius:6px;border:none;background:transparent;color:var(--w4);cursor:pointer;font-family:inherit;letter-spacing:0.3px;transition:all .15s;}
.tab.on{background:var(--s3);color:var(--w2);border:1px solid var(--line2);}
.sel{background:var(--s2);border:1px solid var(--line2);color:var(--w3);padding:5px 10px;border-radius:6px;font-size:11px;font-family:inherit;cursor:pointer;outline:none;}

/* CHART */
.chart-box{position:relative;width:100%;height:190px;}
.chart-box-lg{position:relative;width:100%;height:220px;}
.leg-row{display:flex;gap:18px;margin-bottom:14px;}
.leg{display:flex;align-items:center;gap:6px;font-size:11px;color:var(--w4);}
.leg-dot{width:7px;height:7px;border-radius:50%;flex-shrink:0;}
.fc-desc{font-size:11px;color:var(--w4);margin-bottom:16px;letter-spacing:0.2px;line-height:1.7;}

/* 2COL DETAIL */
.d-two{display:grid;grid-template-columns:1fr 1fr;gap:28px;}
.d-two-col+.d-two-col{padding-left:28px;border-left:1px solid var(--line);}

/* MEMBER GRID */
.pay-toprow{display:flex;align-items:center;justify-content:space-between;margin-bottom:14px;gap:10px;}
.pay-count{font-size:11px;color:var(--w4);}
.pay-count span{color:var(--w2);}
.copy-box{cursor:pointer;background:var(--p-dim);border:1px solid rgba(0,212,160,0.2);border-radius:8px;padding:6px 12px;text-align:right;transition:border-color .2s,background .2s;flex-shrink:0;}
.copy-box.copied{background:rgba(0,212,160,0.2);border-color:rgba(0,212,160,0.5);}
.copy-lbl{font-size:9px;color:rgba(0,212,160,0.6);letter-spacing:0.6px;margin-bottom:2px;}
.copy-num{font-size:12px;font-weight:500;color:var(--p);letter-spacing:0.8px;font-variant-numeric:tabular-nums;}
.mgrid{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;}
.mc{display:flex;flex-direction:column;align-items:center;gap:5px;padding:11px 4px;border-radius:10px;border:1px solid var(--line);min-width:0;}
.mc.paid{background:var(--p-dim);border-color:rgba(0,212,160,0.18);}
.mc.unpaid{background:transparent;}
.mc-av{width:28px;height:28px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:500;flex-shrink:0;}
.mc.paid .mc-av{background:rgba(0,212,160,0.18);color:var(--p);}
.mc.unpaid .mc-av{background:rgba(232,232,234,0.05);color:var(--w4);}
.mc-name{font-size:10px;letter-spacing:-0.2px;max-width:100%;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.mc.paid .mc-name{color:var(--w2);}
.mc.unpaid .mc-name{color:var(--w4);}
.mc-tag{font-size:9px;padding:2px 7px;border-radius:4px;font-weight:500;letter-spacing:0.3px;}
.mc.paid .mc-tag{background:rgba(0,212,160,0.15);color:var(--p);}
.mc.unpaid .mc-tag{background:rgba(232,232,234,0.04);color:var(--w4);}

/* TIMELINE */
.tl{display:flex;flex-direction:column;}
.tl-item{display:flex;gap:12px;padding:11px 0;border-bottom:1px solid var(--line);}
.tl-item:last-child{border-bottom:none;}
.tl-icon{width:26px;height:26px;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:12px;flex-shrink:0;margin-top:1px;}
.tl-icon.in{background:var(--p-dim2);}
.tl-icon.out{background:var(--neg-dim);}
.tl-body{flex:1;min-width:0;}
.tl-row1{display:flex;justify-content:space-between;align-items:baseline;gap:8px;margin-bottom:3px;}
.tl-memo{font-size:12px;color:var(--w2);white-space:nowrap;overflow:hidden;text-overflow:ellipsis;flex:1;}
.tl-amt{font-size:13px;font-weight:500;flex-shrink:0;letter-spacing:-0.2px;font-variant-numeric:tabular-nums;}
.tl-amt.in{color:var(--p);}
.tl-amt.out{color:var(--neg);}
.tl-row2{display:flex;align-items:center;gap:7px;}
.tl-date{font-size:10px;color:var(--w4);font-variant-numeric:tabular-nums;}
.tl-tag{font-size:9px;padding:2px 7px;border-radius:4px;letter-spacing:0.2px;}
.tag-모임비{background:rgba(232,232,234,0.05);color:var(--w4);}
.tag-회비{background:var(--p-dim);color:var(--p);}
.tag-경조사비{background:rgba(232,232,234,0.07);color:var(--w3);}
.tag-강연비{background:rgba(232,232,234,0.07);color:var(--w3);}
.tag-경품비{background:rgba(232,232,234,0.07);color:var(--w3);}
.tag-지각비{background:var(--p-dim);color:rgba(0,212,160,0.6);}
.tag-이자{background:rgba(232,232,234,0.04);color:var(--w4);}
.tag-기타{background:rgba(232,232,234,0.04);color:var(--w4);}

@media(max-width:580px){
  .t-hero-num{font-size:50px;letter-spacing:-1.5px;}
  .t-month-val{font-size:19px;}
  .t-month-row{gap:10px;}
  .t-focus{padding:20px 20px;}
  .t-focus-val{font-size:34px;}
  .d-two{grid-template-columns:1fr;}
  .d-two-col+.d-two-col{padding-left:0;border-left:none;border-top:1px solid var(--line);padding-top:24px;margin-top:24px;}
  .mgrid{grid-template-columns:repeat(3,1fr);}
  .d-head{flex-wrap:wrap;gap:8px;}
}
</style>
</head>
<body>

<nav>
  <span class="nav-logo">사<em>군</em>자</span>
  <div class="nav-right">
    <span class="nav-time" id="nav-time">—</span>
    <button class="nav-btn" id="refresh-btn" onclick="init()">REFRESH</button>
  </div>
</nav>

<div class="page">

  <!-- TIER 1: HERO -->
  <section class="t-hero">
    <div class="t-hero-eyebrow">Fund Balance</div>
    <div class="t-hero-num"><sup>₩</sup><span id="h-bal">0</span></div>
    <div class="t-hero-meta">
      <span id="h-period">—</span>
      <span class="t-hero-dot">·</span>
      <span id="h-net">—</span>
    </div>
  </section>

  <!-- TIER 2: 이번달 흐름 -->
  <section class="t-month">
    <div class="t-month-head">
      <span class="t-month-title">이번달 흐름</span>
      <span class="t-month-period" id="m-period">—</span>
    </div>
    <div class="t-month-row">
      <div>
        <div class="t-month-lbl">입금</div>
        <div class="t-month-val" id="m-in">—</div>
      </div>
      <div>
        <div class="t-month-lbl">지출</div>
        <div class="t-month-val" id="m-out">—</div>
      </div>
      <div>
        <div class="t-month-lbl">순증감</div>
        <div class="t-month-val" id="m-net">—</div>
      </div>
    </div>
  </section>

  <!-- 핵심 운영 지표: 월평균 모임비 (두 번째 강조점) -->
  <section class="t-focus">
    <div class="t-focus-head">
      <div class="t-focus-titlewrap">
        <span class="t-focus-title">월 평균 모임비</span>
        <span class="t-focus-note">경조사비 제외</span>
      </div>
      <div class="tabs" id="avg-toggle">
        <button class="tab" data-n="3" onclick="setAvgPeriod(3,this)">3개월</button>
        <button class="tab on" data-n="6" onclick="setAvgPeriod(6,this)">6개월</button>
        <button class="tab" data-n="12" onclick="setAvgPeriod(12,this)">12개월</button>
      </div>
    </div>
    <div class="t-focus-val" id="h-avg-out">—</div>
    <div class="t-focus-sub" id="h-avg-out-sub">—</div>
    <div class="t-focus-secondary">
      <span class="t-focus-sec-lbl">월 평균 입금</span>
      <span class="t-focus-sec-val" id="h-avg-in">—</span>
      <span class="t-focus-sec-sub" id="h-avg-in-sub">—</span>
    </div>
  </section>

  <!-- TIER 3: 상세 -->
  <section class="t-detail">

    <div class="d-block">
      <div class="d-head"><span class="d-title">누적 총계</span></div>
      <div class="d-row">
        <div class="d-stat">
          <span class="d-stat-lbl">Total In</span>
          <span class="d-stat-val" id="h-in">—</span>
          <span class="d-stat-sub">회비 + 지각비 + 이자</span>
        </div>
        <div class="d-stat">
          <span class="d-stat-lbl">Total Out</span>
          <span class="d-stat-val" id="h-out">—</span>
          <span class="d-stat-sub">모임비 + 경조사비 외</span>
        </div>
      </div>
    </div>

    <div class="d-block">
      <div class="d-head">
        <span class="d-title">Monthly Flow</span>
        <div class="tabs">
          <button class="tab on" onclick="setMode('bar',this)">막대</button>
          <button class="tab" onclick="setMode('line',this)">추세</button>
          <button class="tab" onclick="setMode('cum',this)">누적</button>
        </div>
      </div>
      <div class="leg-row">
        <span class="leg"><span class="leg-dot" style="background:#00D4A0"></span>입금</span>
        <span class="leg"><span class="leg-dot" style="background:#ff5e5e"></span>출금</span>
      </div>
      <div class="chart-box"><canvas id="mainC"></canvas></div>
    </div>

    <div class="d-block">
      <div class="d-head"><span class="d-title">Balance Forecast</span></div>
      <div id="fc-desc" class="fc-desc"></div>
      <div class="chart-box-lg"><canvas id="fcC"></canvas></div>
    </div>

    <div class="d-block">
      <div class="d-two">
        <div class="d-two-col">
          <div class="d-head">
            <span class="d-title">Monthly Payment</span>
            <select id="pay-select" onchange="onPaySelect(this.value)" class="sel"></select>
          </div>
          <div class="pay-toprow">
            <span class="pay-count" id="pay-count">—</span>
            <div id="copy-acc" onclick="copyAccount()" title="클릭하면 계좌번호 복사" class="copy-box">
              <div class="copy-lbl">카카오뱅크 · 조현민</div>
              <div id="acc-num" class="copy-num">7979-0864-002</div>
            </div>
          </div>
          <div class="mgrid" id="mgrid"></div>
        </div>
        <div class="d-two-col">
          <div class="d-head"><span class="d-title">Recent Transactions</span></div>
          <div class="tl" id="tl"></div>
        </div>
      </div>
    </div>

  </section>

</div>

<script>
const fmtFull = n => Math.round(Math.abs(n)).toLocaleString('ko-KR');
const fmtK = n => {
  const a = Math.round(Math.abs(n));
  if(a >= 100000000){
    const eok = Math.floor(a/100000000);
    const man = Math.round((a%100000000)/10000);
    return man>0 ? `${eok}억 ${man.toLocaleString('ko-KR')}만` : `${eok}억`;
  }
  if(a >= 10000) return Math.round(a/10000).toLocaleString('ko-KR')+'만';
  return a.toLocaleString('ko-KR');
};

function animateNum(el, target, fmt, dur=700){
  if(el.__timer) clearTimeout(el.__timer);
  const t0=Date.now();
  function step(){
    const p=Math.min((Date.now()-t0)/dur,1);
    const e=1-Math.pow(1-p,3);
    el.textContent=fmt(target*e);
    if(p<1) el.__timer=setTimeout(step,16);
  }
  step();
}

let D=null, MC=null, FC=null, mode='bar', avgPeriod=6;

async function init(){
  const btn=document.getElementById('refresh-btn');
  btn.disabled=true; btn.textContent='LOADING…'; btn.style.opacity=0.5;
  document.getElementById('nav-time').textContent='…';
  try{
    const r=await fetch('/api/data'); D=await r.json(); render();
  } finally {
    btn.disabled=false; btn.textContent='REFRESH'; btn.style.opacity=1;
  }
}

function copyAccount(){
  try{ navigator.clipboard.writeText('79790864002'); }catch(e){}
  const box=document.getElementById('copy-acc'), num=document.getElementById('acc-num');
  const orig=num.textContent;
  num.textContent='✓ 복사완료';
  box.classList.add('copied');
  clearTimeout(window.__copyTimer);
  window.__copyTimer=setTimeout(()=>{
    num.textContent=orig;
    box.classList.remove('copied');
  },1500);
}

function render(){
  const s=D.summary;
  document.getElementById('nav-time').textContent=s.updated;
  animateNum(document.getElementById('h-bal'), s.balance, fmtFull);

  const real=D.monthly.filter(m=>m.income>0||m.expense>0);
  document.getElementById('h-period').textContent=real.length?`${real[0].month} — ${real[real.length-1].month}`:'';
  const net=s.total_in-s.total_out;
  const netEl=document.getElementById('h-net');
  netEl.classList.toggle('neg', net<0);
  netEl.textContent=(net>=0?'▲ +':'▼ ')+fmtK(net)+'원 순증';

  animateNum(document.getElementById('h-in'), s.total_in, v=>fmtK(v)+'원');
  animateNum(document.getElementById('h-out'), s.total_out, v=>fmtK(v)+'원');

  const cur=real[real.length-1];
  if(cur){
    const netCur=cur.income-cur.expense;
    document.getElementById('m-period').textContent=cur.month;
    animateNum(document.getElementById('m-in'), cur.income, v=>fmtK(v)+'원');
    animateNum(document.getElementById('m-out'), cur.expense, v=>fmtK(v)+'원');
    const mNet=document.getElementById('m-net');
    mNet.classList.remove('pos','neg');
    mNet.classList.add(netCur>=0?'pos':'neg');
    animateNum(mNet, netCur, v=>(v>=0?'+':'-')+fmtK(v)+'원');
  }

  updateAvgCards();
  buildMain(); buildFc(); buildPaySelect(); buildMembers(); buildTL();
}

function setAvgPeriod(n,btn){
  avgPeriod=n;
  document.querySelectorAll('#avg-toggle .tab').forEach(t=>t.classList.remove('on'));
  btn.classList.add('on');
  updateAvgCards();
}

function updateAvgCards(){
  const inc=D.summary.income_stats[avgPeriod]||{mean:0,median:0};
  const exp=D.summary.meeting_expense_stats[avgPeriod]||{mean:0,median:0};
  document.getElementById('h-avg-in').textContent=fmtK(inc.mean)+'원';
  document.getElementById('h-avg-in-sub').textContent=`최근 ${avgPeriod}개월 평균 · 보통 ${fmtK(inc.median)}원`;
  document.getElementById('h-avg-out').textContent=fmtK(exp.mean)+'원';
  document.getElementById('h-avg-out-sub').textContent=`최근 ${avgPeriod}개월 평균 · 보통 ${fmtK(exp.median)}원`;
}

function buildMain(){
  const real=D.monthly.filter(m=>m.income>0||m.expense>0);
  const labels=real.map(m=>m.month);
  let ds;
  if(mode==='cum'){
    ds=[{type:'line',data:real.map(m=>m.cumulative),borderColor:'rgba(255,255,255,0.5)',backgroundColor:'rgba(255,255,255,0.03)',fill:true,tension:0.4,pointRadius:0,borderWidth:1.5,borderDash:[4,3]}];
  } else {
    ds=[
      {type:mode,label:'입금',data:real.map(m=>m.income),backgroundColor:mode==='bar'?'rgba(0,212,160,0.65)':'transparent',borderColor:'#00D4A0',borderRadius:mode==='bar'?3:0,borderWidth:1.5,tension:0.4,pointRadius:0,fill:false},
      {type:mode,label:'출금',data:real.map(m=>m.expense),backgroundColor:mode==='bar'?'rgba(255,94,94,0.55)':'transparent',borderColor:'#ff5e5e',borderRadius:mode==='bar'?3:0,borderWidth:1.5,tension:0.4,pointRadius:0,fill:false},
    ];
  }
  if(MC){
    MC.data.labels=labels;
    MC.data.datasets=ds;
    MC.update();
  } else {
    MC=new Chart(document.getElementById('mainC'),{
      data:{labels,datasets:ds},
      options:{responsive:true,maintainAspectRatio:false,
        animation:{duration:500,easing:'easeOutQuart'},
        plugins:{legend:{display:false},tooltip:{callbacks:{label:c=>fmtK(c.raw)+'원'}}},
        scales:{
          x:{ticks:{color:'rgba(232,232,234,0.28)',font:{size:10},maxRotation:45,autoSkip:true,maxTicksLimit:10},grid:{display:false},border:{display:false}},
          y:{ticks:{color:'rgba(232,232,234,0.28)',font:{size:10},callback:v=>fmtK(v)},grid:{color:'rgba(232,232,234,0.05)'},border:{display:false}}
        }}
    });
  }
}

function iqrFilter(arr){
  const sorted=[...arr].sort((a,b)=>a-b);
  const q1=sorted[Math.floor(sorted.length*0.25)];
  const q3=sorted[Math.floor(sorted.length*0.75)];
  const iqr=q3-q1;
  const upper=q3+1.5*iqr;
  return arr.filter(v=>v<=upper);
}

function buildFc(){
  const real=D.monthly.filter(m=>m.income>0||m.expense>0);
  const li=real.filter(m=>m.income>0).slice(-6);
  const ai=Math.round(li.reduce((a,m)=>a+m.income,0)/Math.max(li.length,1));
  const lo=real.filter(m=>m.expense>0).slice(-14);
  const loVals=lo.map(m=>m.expense);
  const loFiltered=iqrFilter(loVals);
  const ao=Math.round(loFiltered.reduce((a,v)=>a+v,0)/Math.max(loFiltered.length,1));
  const excluded=loVals.length-loFiltered.length;
  const cb=D.summary.balance;

  // 단일 트렌드 라인 + 낙관(지출-20%)~비관(지출+20%) 범위를 음영 밴드로 표현.
  // 실제 입금/출금(초록·빨강)과 구분되는 블루 색조 유지.
  const hist=real.slice(-5);
  const lastM=hist[hist.length-1]?.month||'';
  const lm=lastM.match(/(\d+)년 (\d+)월/);
  let baseYY=lm?parseInt(lm[1]):26, baseMM=lm?parseInt(lm[2]):5;
  const fcLabels=[];
  for(let i=1;i<=12;i++){
    let mm=baseMM+i,yy=baseYY+Math.floor((mm-1)/12);
    mm=((mm-1)%12)+1;
    fcLabels.push(`${yy%100}년 ${mm}월`);
  }
  const allL=[...hist.map(m=>m.month),...fcLabels];
  const histConnect=hist[hist.length-1].cumulative;

  function project(outAdj){
    const pts=[]; let bal=cb;
    for(let i=0;i<12;i++){ bal += ai-Math.round(ao*outAdj); pts.push(Math.round(bal)); }
    return pts;
  }
  const trendPts=project(1.0);
  const optPts=project(0.8);   // 지출 20% 적게 → 낙관(잔액 상단)
  const pessPts=project(1.2);  // 지출 20% 많이 → 비관(잔액 하단)
  const nullsHist=new Array(4).fill(null);

  const datasets=[
    {label:'실제', data:[...hist.map(m=>m.cumulative), ...new Array(12).fill(null)],
     borderColor:'rgba(232,232,234,0.45)', backgroundColor:'transparent',
     fill:false, tension:0.35, pointRadius:0, borderWidth:2, order:1},
    {label:'낙관', data:[...nullsHist, histConnect, ...optPts],
     borderColor:'transparent', backgroundColor:'transparent',
     fill:false, tension:0.35, pointRadius:0, borderWidth:0, order:4},
    {label:'비관', data:[...nullsHist, histConnect, ...pessPts],
     borderColor:'transparent', backgroundColor:'rgba(77,163,255,0.09)',
     fill:'-1', tension:0.35, pointRadius:0, borderWidth:0, order:3},
    {label:'현재 추세', data:[...nullsHist, histConnect, ...trendPts],
     borderColor:'#4da3ff', backgroundColor:'transparent',
     fill:false, tension:0.35, pointRadius:0, borderWidth:2, borderDash:[5,4], order:2},
  ];

  if(FC){
    FC.data.labels=allL;
    FC.data.datasets=datasets;
    FC.update();
  } else {
    FC=new Chart(document.getElementById('fcC'),{
      type:'line',
      data:{labels:allL,datasets},
      options:{responsive:true,maintainAspectRatio:false,
        animation:{duration:500,easing:'easeOutQuart'},
        plugins:{
          legend:{display:false},
          tooltip:{
            mode:'index',intersect:false,
            filter:c=>c.dataset.label==='실제'||c.dataset.label==='현재 추세',
            callbacks:{ label:c=>c.raw!=null?` ${c.dataset.label}: ${fmtK(c.raw)}원`:null }
          }
        },
        scales:{
          x:{ticks:{color:'rgba(232,232,234,0.28)',font:{size:10},maxRotation:0,autoSkip:true,maxTicksLimit:6},grid:{display:false},border:{display:false}},
          y:{ticks:{color:'rgba(232,232,234,0.28)',font:{size:10},callback:v=>fmtK(v)},grid:{color:'rgba(232,232,234,0.05)'},border:{display:false}}
        }}
    });
  }

  const bal12Trend=trendPts[11], bal12Opt=optPts[11], bal12Pess=pessPts[11];
  const desc=document.getElementById('fc-desc');
  if(desc) desc.innerHTML=
    `입금 기준 최근 6개월 평균 <b style="color:var(--w2)">${fmtK(ai)}원</b> · `+
    `지출 기준 최근 12개월 평균 <b style="color:var(--w2)">${fmtK(ao)}원</b>(비경상 ${excluded}개월 제외)<br>`+
    `12개월 후 예상 잔액 <b style="color:#4da3ff">${fmtK(bal12Trend)}원</b> `+
    `<span style="color:var(--w4)">(범위 ${fmtK(bal12Pess)}~${fmtK(bal12Opt)}원)</span>`;
}

function buildPaySelect(){
  const sel=document.getElementById('pay-select');
  sel.innerHTML='';
  // 최신 월이 맨 위로
  [...D.pay_months].reverse().forEach(ym=>{
    const opt=document.createElement('option');
    opt.value=ym; opt.textContent=ym;
    if(ym===D.summary.latest_ym) opt.selected=true;
    sel.appendChild(opt);
  });
}

function buildMembers(ym){
  ym=ym||D.summary.latest_ym;
  const ms=D.payment_by_month[ym]||[];
  const pc=ms.filter(m=>m.paid).length;
  document.getElementById('pay-count').innerHTML=`<span style="color:#00D4A0;font-weight:500">${pc}명</span> / ${ms.length}명 납부완료`;
  const g=document.getElementById('mgrid'); g.innerHTML='';
  // 납부자 먼저, 미납자 뒤에
  const sorted=[...ms].sort((a,b)=>b.paid-a.paid);
  sorted.forEach(m=>{
    const c=m.paid?'paid':'unpaid';
    g.innerHTML+=`<div class="mc ${c}">
      <div class="mc-av">${m.name[0]}</div>
      <div class="mc-name">${m.name}</div>
      <span class="mc-tag">${m.paid?'납부':'미납'}</span>
    </div>`;
  });
}

function onPaySelect(ym){ buildMembers(ym); }

function buildTL(){
  const t=document.getElementById('tl'); t.innerHTML='';
  D.recent.forEach(r=>{
    const cls=r.is_income?'in':'out',sign=r.is_income?'+':'-';
    const icon=r.is_income?'↑':'↓';
    t.innerHTML+=`<div class="tl-item">
      <div class="tl-icon ${cls}">${icon}</div>
      <div class="tl-body">
        <div class="tl-row1">
          <span class="tl-memo">${r.memo||'—'}</span>
          <span class="tl-amt ${cls}">${sign}${fmtFull(r.amount)}원</span>
        </div>
        <div class="tl-row2">
          <span class="tl-date">${r.date}</span>
          <span class="tl-tag tag-${r.type}">${r.type}</span>
        </div>
      </div>
    </div>`;
  });
}

function setMode(m,b){mode=m;document.querySelectorAll('.tab').forEach(t=>t.classList.remove('on'));b.classList.add('on');buildMain();}


init();
</script>
</body>
</html>'''

class Handler(BaseHTTPRequestHandler):
    def log_message(self,f,*a): pass
    def do_HEAD(self):
        p=urlparse(self.path).path
        if p=='/api/data':
            self.send_response(200);self.send_header('Content-Type','application/json;charset=utf-8');self.end_headers()
        else:
            self.send_response(200);self.send_header('Content-Type','text/html;charset=utf-8');self.end_headers()
    def do_GET(self):
        p=urlparse(self.path).path
        if p=='/api/data':
            try:
                body=json.dumps(build_data(),ensure_ascii=False).encode('utf-8')
                self.send_response(200);self.send_header('Content-Type','application/json;charset=utf-8');self.send_header('Content-Length',len(body));self.end_headers();self.wfile.write(body)
            except Exception as e:
                import traceback
                body=json.dumps({'error':str(e),'trace':traceback.format_exc()}).encode()
                self.send_response(500);self.send_header('Content-Type','application/json');self.end_headers();self.wfile.write(body)
        else:
            body=HTML.encode('utf-8')
            self.send_response(200);self.send_header('Content-Type','text/html;charset=utf-8');self.send_header('Content-Length',len(body));self.end_headers();self.wfile.write(body)

if __name__=='__main__':
    import socket, webbrowser, threading, time, sys

    # Render는 PORT 환경변수로 포트를 지정함
    env_port = os.environ.get('PORT')
    if env_port:
        port = int(env_port)
        server = ThreadingHTTPServer(('0.0.0.0', port), Handler)
        print(f'사군자 대시보드 실행 중: http://0.0.0.0:{port}')
        try: server.serve_forever()
        except KeyboardInterrupt: print('\n서버 종료.')
    else:
        # 로컬 실행: 빈 포트 자동 탐색 + 브라우저 자동 열기
        def find_free_port(start=8000):
            for p in range(start, start+10):
                try:
                    s=socket.socket(); s.bind(('',p)); s.close(); return p
                except: continue
            return start

        start_port = 8000
        for i, arg in enumerate(sys.argv):
            if arg == '--port' and i+1 < len(sys.argv):
                start_port = int(sys.argv[i+1])

        port = find_free_port(start_port)
        url  = f'http://localhost:{port}'

        try:
            server = ThreadingHTTPServer(('0.0.0.0', port), Handler)
        except OSError:
            print(f'포트 {port} 사용 중. 브라우저에서 {url} 을 열어보세요.')
            webbrowser.open(url); exit()

        threading.Thread(target=lambda: (time.sleep(1.2), webbrowser.open(url)), daemon=True).start()

        print(f"""
  ╔══════════════════════════════════════╗
  ║   사군자 회비 대시보드 실행 중        ║
  ╠══════════════════════════════════════╣
  ║  주소:  {url:<28} ║
  ║  종료:  Ctrl+C  또는 창 닫기         ║
  ╚══════════════════════════════════════╝
""")
        try: server.serve_forever()
        except KeyboardInterrupt: print('\n  서버 종료.')
