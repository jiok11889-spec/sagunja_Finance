"""
카카오뱅크 거래내역 CSV → 사군자 엑셀 자동 업데이트
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import sys
import os
import re

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', '사군자.xlsx')

# 카카오뱅크 분류 매핑 규칙 (메모/거래내용 키워드 → 분류1, 분류2)
CATEGORY_RULES = [
    # 입금
    ({'회비', '월간회비', '회비납부'}, '입금', '회비'),
    ({'지각비'}, '입금', '지각비'),
    ({'이자'}, '입금', '이자'),
    # 출금
    ({'모임', '식사', '저녁', '점심', '식당', '파티룸', '바', '로프트', '라운지', '레스토랑', '카페'}, '출금', '모임비'),
    ({'강연', 'AI강의', '특강', '강의'}, '출금', '강연비'),
    ({'축의금', '조의금', '생일', '화환', '경조사'}, '출금', '경조사비'),
    ({'경품'}, '출금', '경품비'),
]

def classify_transaction(row):
    """거래내역 한 행을 분류1, 분류2로 분류"""
    text = ' '.join([
        str(row.get('거래내용', '')),
        str(row.get('메모', ''))
    ]).lower()

    for keywords, cat1, cat2 in CATEGORY_RULES:
        if any(kw in text for kw in keywords):
            return cat1, cat2

    # 기본 처리
    if row.get('거래유형') == '입금':
        return '입금', '기타'
    else:
        return '출금', '기타'


def parse_kakao_csv(csv_path):
    """카카오뱅크 CSV 파싱"""
    # 카카오뱅크 CSV는 인코딩이 cp949 또는 utf-8-sig
    for enc in ['utf-8-sig', 'cp949', 'utf-8']:
        try:
            df = pd.read_csv(csv_path, encoding=enc)
            break
        except Exception:
            continue

    # 날짜 파싱
    df['거래일시'] = pd.to_datetime(df['거래일시'])
    df['연월'] = df['거래일시'].apply(lambda d: f"{str(d.year)[2:]}년 {d.month}월")
    df['일자'] = df['거래일시'].dt.date

    # 요일 한국어
    weekday_map = {0: '월', 1: '화', 2: '수', 3: '목', 4: '금', 5: '토', 6: '일'}
    df['요일'] = df['거래일시'].dt.weekday.map(weekday_map)

    # 분류
    df[['분류1', '분류2']] = df.apply(
        lambda r: pd.Series(classify_transaction(r)), axis=1
    )

    # 차변/대변
    df['차변'] = df.apply(lambda r: abs(r['거래금액']) if r['거래유형'] == '출금' else None, axis=1)
    df['대변'] = df.apply(lambda r: abs(r['거래금액']) if r['거래유형'] == '입금' else None, axis=1)
    df['대변-차변'] = df.apply(
        lambda r: r['대변'] if pd.notna(r.get('대변')) else -r['차변'], axis=1
    )
    df['비고'] = df.apply(lambda r: str(r.get('메모', '') or r.get('거래내용', '')), axis=1)

    return df


def get_existing_dates(wb):
    """원장 시트에서 이미 존재하는 (일자, 비고) set 반환"""
    ws = wb['원장']
    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_val = row[2]
        memo_val = row[8]
        if date_val and date_val != '입력':
            if hasattr(date_val, 'date'):
                existing.add((str(date_val.date()), str(memo_val or '')))
            else:
                existing.add((str(date_val)[:10], str(memo_val or '')))
    return existing


def append_to_wonjang(wb, new_rows):
    """원장 시트에 새 거래 추가"""
    ws = wb['원장']

    # 마지막 데이터 행 찾기 (헤더 row=1, 수식설명 row=2)
    last_row = ws.max_row
    while last_row > 2 and ws.cell(last_row, 3).value is None:
        last_row -= 1

    added = 0
    for _, r in new_rows.iterrows():
        last_row += 1
        ws.cell(last_row, 1).value = r['연월']
        ws.cell(last_row, 2).value = r['요일']
        ws.cell(last_row, 3).value = r['일자']
        ws.cell(last_row, 4).value = r['분류1']
        ws.cell(last_row, 5).value = r['분류2']
        ws.cell(last_row, 6).value = r['차변'] if pd.notna(r['차변']) else None
        ws.cell(last_row, 7).value = r['대변'] if pd.notna(r['대변']) else None
        # 대변-차변 값 직접 계산
        d = float(r['대변']) if pd.notna(r['대변']) and r['대변'] else 0
        c = float(r['차변']) if pd.notna(r['차변']) and r['차변'] else 0
        ws.cell(last_row, 8).value = d - c
        ws.cell(last_row, 9).value = r['비고']
        added += 1

    return added


def update_excel(csv_path):
    print(f"\n📂 CSV 파일 읽는 중: {csv_path}")
    df = parse_kakao_csv(csv_path)
    print(f"   → {len(df)}건 파싱 완료")

    print(f"\n📊 엑셀 파일 열기: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    existing = get_existing_dates(wb)
    print(f"   → 기존 원장 거래 {len(existing)}건")

    # 중복 제거
    def is_new(r):
        date_str = str(r['일자'])
        memo_str = str(r['비고'])
        return (date_str, memo_str) not in existing

    new_df = df[df.apply(is_new, axis=1)].copy()
    new_df = new_df.sort_values('거래일시')

    if len(new_df) == 0:
        print("\n✅ 새로운 거래내역이 없습니다. 엑셀이 이미 최신입니다.")
        return

    print(f"\n➕ 새 거래 {len(new_df)}건 추가 예정:")
    for _, r in new_df.iterrows():
        amt = f"+{int(r['대변']):,}" if pd.notna(r['대변']) else f"-{int(r['차변']):,}"
        print(f"   {r['연월']} {r['일자']} [{r['분류2']}] {amt}원  {r['비고']}")

    added = append_to_wonjang(wb, new_df)
    wb.save(EXCEL_PATH)
    print(f"\n✅ {added}건 추가 완료! 파일 저장됨: {EXCEL_PATH}")
    print("\n💡 엑셀을 열어 '회비입출내역', '증감분석' 시트의 수식을 확인해 주세요.")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("사용법: python update_excel.py <카카오뱅크_거래내역.csv>")
        print("예시:  python update_excel.py ../sample/카카오뱅크_거래내역_샘플.csv")
        sys.exit(1)

    update_excel(sys.argv[1])
